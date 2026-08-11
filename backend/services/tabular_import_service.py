from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


class ExcelImportError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedExcelTable:
    sheet_name: str
    header_row: int
    headers: tuple[str | None, ...]
    records: list[dict[str, Any]]


def clean_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_header(value: Any, aliases: dict[str, Iterable[str]]) -> str | None:
    raw = clean_cell_text(value).lower().replace(" ", "").replace("-", "_")
    if not raw:
        return None
    for field, values in aliases.items():
        normalized = {str(alias).lower().replace(" ", "").replace("-", "_") for alias in values}
        if raw in normalized:
            return field
    return None


def _cell_text(cell: Cell) -> str:
    value = cell.value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number_format = str(cell.number_format or "").strip()
        if re.fullmatch(r"0+", number_format):
            try:
                return str(int(value)).zfill(len(number_format))
            except (TypeError, ValueError):
                pass
    return clean_cell_text(value)


def parse_excel_table(
    content: bytes,
    *,
    aliases: dict[str, Iterable[str]],
    required_field: str,
    allowed_fields: Iterable[str] | None = None,
    max_bytes: int = 15 * 1024 * 1024,
) -> ParsedExcelTable:
    if not content:
        raise ExcelImportError("Excel文件为空")
    if len(content) > max_bytes:
        raise ExcelImportError(f"Excel文件不能超过 {max_bytes // (1024 * 1024)}MB")
    workbook = None
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    except Exception as error:
        raise ExcelImportError(f"Excel文件无法解析：{error}")

    allowed = set(allowed_fields or aliases.keys())
    candidates: list[tuple[int, int, int, Any, list[str | None], int]] = []
    discovered: list[str] = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets):
            sheet_fields: list[str] = []
            for row_index, row in enumerate(sheet.iter_rows(), start=1):
                values = [_cell_text(cell) for cell in row]
                if not any(values):
                    continue
                mapped = [normalize_header(value, aliases) for value in values]
                recognized = {field for field in mapped if field in allowed}
                if required_field in recognized:
                    visible_priority = 1 if sheet.sheet_state == "visible" else 0
                    candidates.append((len(recognized), visible_priority, -sheet_index, sheet, mapped, row_index))
                if len(sheet_fields) < 10:
                    sheet_fields.extend(value[:24] for value in values if value and value not in sheet_fields)
            if sheet_fields:
                discovered.append(f"{sheet.title}：{'、'.join(sheet_fields[:10])}")

        if not candidates:
            details = "；".join(discovered[:4]) or "工作簿中没有可读取内容"
            raise ExcelImportError(f"未找到“学号”字段。已扫描字段：{details}")

        _, _, _, sheet, headers, header_row = max(candidates, key=lambda item: item[:3])
        records: list[dict[str, Any]] = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
            values = [_cell_text(cell) for cell in row]
            if not any(values):
                continue
            record: dict[str, Any] = {"row_number": row_index, "sheet_name": sheet.title}
            for column_index, field in enumerate(headers):
                if field and field in allowed and field not in record:
                    record[field] = values[column_index] if column_index < len(values) else ""
            if any(clean_cell_text(record.get(field)) for field in allowed):
                records.append(record)
        if not records:
            raise ExcelImportError(f"工作表“{sheet.title}”的学号表头下没有可导入数据")
        return ParsedExcelTable(
            sheet_name=sheet.title,
            header_row=header_row,
            headers=tuple(headers),
            records=records,
        )
    finally:
        if workbook is not None:
            workbook.close()
